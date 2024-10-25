import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# 祝日リストを作成（例として2024年の一部の祝日）
holidays = ['2024-01-01', '2024-01-08', '2024-02-11', '2024-02-23', '2024-03-20',
            '2024-04-29', '2024-05-03', '2024-05-04', '2024-05-05', '2024-05-06',
            '2024-07-15', '2024-08-11', '2024-09-16', '2024-09-23', '2024-10-14',
            '2024-11-03', '2024-11-23']

holidays = [datetime.strptime(date, '%Y-%m-%d').date() for date in holidays]

# 開始日と終了日を設定
start_date = datetime(2024, 5, 17)
end_date = datetime(2024, 6, 5)

# 日付リストを作成
dates = pd.date_range(start=start_date, end=end_date).tolist()

# スケジュール表の初期データフレームを作成
items = ['項目1', '項目2', '項目3']  # 項目のリストを設定
columns = ['項目', '合計台数'] + [f'{date.strftime("%m/%d")}' for date in dates for _ in range(2)]
df = pd.DataFrame(columns=columns)

# 項目をデータフレームに追加
df['項目'] = items

# データフレームをExcelに書き出し
excel_path = 'kitting_schedule.xlsx'
df.to_excel(excel_path, sheet_name='スケジュール', index=False)

# Excelファイルを開いて書式設定
wb = openpyxl.load_workbook(excel_path)
ws = wb['スケジュール']

# フォント設定
font = Font(name='Meiryo UI')

# カラー設定
header_fill = PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid")
gray_fill = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ヘッダー行に日付を設定し、セルを結合
for i, date in enumerate(dates):
    col1 = 3 + i * 2
    col2 = col1 + 1
    ws.merge_cells(start_row=1, start_column=col1, end_row=1, end_column=col2)
    ws.cell(row=1, column=col1).value = date.strftime('%m/%d')
    ws.cell(row=1, column=col1).font = font
    ws.cell(row=1, column=col1).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=col1).fill = header_fill
    ws.cell(row=2, column=col1).value = '予定'
    ws.cell(row=2, column=col2).value = '実績'
    ws.cell(row=2, column=col1).font = font
    ws.cell(row=2, column=col2).font = font
    ws.cell(row=2, column=col1).fill = header_fill
    ws.cell(row=2, column=col2).fill = header_fill

    # 土日および祝日のグレーアウト
    if date.weekday() >= 5 or date in holidays:
        ws.cell(row=1, column=col1).fill = gray_fill
        ws.cell(row=1, column=col2).fill = gray_fill
        ws.cell(row=2, column=col1).fill = gray_fill
        ws.cell(row=2, column=col2).fill = gray_fill

# 項目行の下に作業予定人数の行を追加
ws.insert_rows(2)
for i, date in enumerate(dates):
    col1 = 3 + i * 2
    col2 = col1 + 1
    ws.merge_cells(start_row=2, start_column=col1, end_row=2, end_column=col2)
    ws.cell(row=2, column=col1).value = '作業予定人数'
    ws.cell(row=2, column=col1).font = font
    ws.cell(row=2, column=col1).alignment = Alignment(horizontal='center')
    ws.cell(row=2, column=col1).fill = header_fill

# 合計台数の列ヘッダーを設定
ws.cell(row=1, column=2).value = '合計台数'
ws.cell(row=1, column=2).font = font
ws.cell(row=1, column=2).fill = header_fill
ws.cell(row=2, column=2).fill = header_fill

# 進捗台数と進捗率の列を追加
progress_qty_col = len(columns) + 1
progress_rate_col = len(columns) + 2
ws.cell(row=1, column=progress_qty_col).value = '進捗台数'
ws.cell(row=1, column=progress_qty_col).font = font
ws.cell(row=1, column=progress_qty_col).fill = header_fill
ws.cell(row=1, column=progress_rate_col).value = '進捗率（％）'
ws.cell(row=1, column=progress_rate_col).font = font
ws.cell(row=1, column=progress_rate_col).fill = header_fill

# 各行の進捗計算式を追加
for row in range(4, len(items) + 4):
    total_col = 'B'  # 合計台数の列
    progress_qty_cell = ws.cell(row=row, column=progress_qty_col)
    progress_rate_cell = ws.cell(row=row, column=progress_rate_col)
    actuals = [f'{get_column_letter(4 + i * 2)}{row}' for i in range(len(dates))]
    progress_qty_cell.value = f'=SUM({",".join(actuals)})'
    progress_rate_cell.value = f'=IF({total_col}{row}>0,{progress_qty_cell.coordinate}/{total_col}{row}*100,0)'
    progress_rate_cell.number_format = '0.00%'

# 進捗台数のフォーマット設定
for row in range(4, len(items) + 4):
    progress_qty_cell = ws.cell(row=row, column=progress_qty_col)
    progress_qty_cell.number_format = '#,##0"台"'

# フォントと格子をすべてのセルに設定
for row in ws.iter_rows():
    for cell in row:
        cell.font = font
        cell.border = border

wb.save(excel_path)